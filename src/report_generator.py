from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def format_excel(file_path):
    wb = load_workbook(file_path)

    header_fill = PatternFill("solid", fgColor="D9E1F2")
    bold = Font(bold=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.value = str(cell.value).title()
            cell.font = bold
            cell.fill = header_fill
            cell.border = border

        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border

        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = max_len + 4

        if "Amount" in [cell.value for cell in sheet[1]]:
            amount_col = [cell.column for cell in sheet[1] if cell.value == "Amount"][0]
            for cell in sheet.iter_rows(min_row=2, min_col=amount_col, max_col=amount_col):
                for c in cell:
                    c.number_format = '₹#,##0.00'

    wb.save(file_path)
